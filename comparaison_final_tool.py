# fichier : comparaison_gradio_updated.py
import unidecode
import gradio as gr
import pandas as pd
import tempfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from openpyxl.utils import get_column_letter
from matplotlib import pyplot as plt

def auto_adjust_column_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

def compute_unique_ids(file, sheet_name, header_row):
    if file is None or sheet_name is None or header_row is None:
        return pd.DataFrame()
    try:
        df_raw = pd.read_excel(file.name, sheet_name=sheet_name, header=None)
        header_row_index = int(header_row) - 1
        new_header = df_raw.iloc[header_row_index]
        df = df_raw.iloc[header_row_index + 1:].copy()
        df.columns = new_header
        df = df.dropna(axis=1, how='all')
        df = df.dropna(how='all')
        df.columns = df.columns.astype(str).str.strip()
        total_rows = len(df)
        unique_counts = df.nunique(dropna=True)
        result = pd.DataFrame({
            "Colonne": unique_counts.index.astype(str),
            "Nb_valeurs_uniques": unique_counts.values
        }).sort_values(by="Nb_valeurs_uniques", ascending=False)
        top5 = result.head(5).copy()
        top5["Total_lignes"] = total_rows
        top5["% unicité"] = (top5["Nb_valeurs_uniques"] / total_rows * 100).round(2)
        return top5
    except Exception as e:
        return pd.DataFrame({"Erreur": [str(e)]})

def normalize_dataframe(df):
    def clean_value(v):
        if pd.isna(v):
            return ""
        if isinstance(v, (int, float, np.number)):
            try:
                return float(v)
            except:
                return ""
        v = str(v)
        v = v.replace("\xa0", " ")
        v = v.replace("\n", " ")
        v = v.replace("\r", " ")
        v = " ".join(v.split())
        v = v.strip().lower()
        return v
    for col in df.columns:
        if col != "merge_key" and col != "occurrence":
            df[col] = df[col].apply(clean_value)
    return df

def safe_compare(v1, v2, decimals =4 ):
    # Deux NaN / vide
    if (pd.isna(v1) or v1 == "") and (pd.isna(v2) or v2 == ""):
        return True  # but we'll treat empties separately later; keep True to avoid float conversion errors
    # Un seul NaN / vide
    if (pd.isna(v1) or v1 == "") or (pd.isna(v2) or v2 == ""):
        return False
    # Si numérique -> arrondi contrôlé
    try:
        f1 = float(v1)
        f2 = float(v2)
        return round(f1, decimals) == round(f2, decimals)
    except:
        pass
    # Sinon texte
    return str(v1).strip() == str(v2).strip()

def normalize_colname(col):
    return unidecode.unidecode(str(col)).strip().lower()

def filter_named_columns(columns):
    filtered = [c for c in columns if str(c).strip() != "" and not str(c).lower().startswith("unnamed")]
    return filtered

def compare_excels(file1, sheet1, header1, col1, file2, sheet2, header2, col2, output_name):
    try:
        # lecture
        df1_raw = pd.read_excel(file1.name, sheet_name=sheet1, header=int(header1)-1, dtype=object)
        df2_raw = pd.read_excel(file2.name, sheet_name=sheet2, header=int(header2)-1, dtype=object)

        # normaliser noms colonnes
        df1_raw.columns = [normalize_colname(c) for c in df1_raw.columns]
        df2_raw.columns = [normalize_colname(c) for c in df2_raw.columns]

        # filrer colonnes sans nom
        df1_raw = df1_raw.loc[:, filter_named_columns(df1_raw.columns)]
        df2_raw = df2_raw.loc[:, filter_named_columns(df2_raw.columns)]

        col1_norm = normalize_colname(col1)
        col2_norm = normalize_colname(col2)

        if col1_norm not in df1_raw.columns:
            return None, f"❌ Erreur : clé '{col1}' introuvable dans le fichier 1", None, pd.DataFrame()
        if col2_norm not in df2_raw.columns:
            return None, f"❌ Erreur : clé '{col2}' introuvable dans le fichier 2", None, pd.DataFrame()

        df1_raw = df1_raw.rename(columns={col1_norm: "merge_key"})
        df2_raw = df2_raw.rename(columns={col2_norm: "merge_key"})
        df1 = df1_raw.copy()
        df2 = df2_raw.copy()

        # Trier
        df1 = df1.sort_values(by=["merge_key"] + [c for c in df1.columns if c != "merge_key"])
        df2 = df2.sort_values(by=["merge_key"] + [c for c in df2.columns if c != "merge_key"])

        # occurrence
        df1["occurrence"] = df1.groupby("merge_key").cumcount()
        df2["occurrence"] = df2.groupby("merge_key").cumcount()

        # normaliser valeurs
        df1 = normalize_dataframe(df1)
        df2 = normalize_dataframe(df2)

        merged = pd.merge(
            df1,
            df2,
            on=["merge_key", "occurrence"],
            how="outer",
            suffixes=('_1', '_2')
        )

        common_cols = [
            c for c in df1.columns
            if c not in ["merge_key", "occurrence"] and c in df2.columns
        ]

        if not common_cols:
            return None, "❌ Erreur : aucune colonne commune à comparer.", None, pd.DataFrame()

        # Construire résultat (paires)
        result = merged[['merge_key']].copy()
        for col in common_cols:
            result[f"{col}_1"] = merged.get(f"{col}_1")
            result[f"{col}_2"] = merged.get(f"{col}_2")

        # Préparer Excel coloré
        wb = Workbook()
        ws = wb.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        for r in dataframe_to_rows(result, index=False, header=True):
            ws.append(r)

        # compteurs globaux
        total_rows = len(merged)
        nb_cols = len(common_cols)

        # stats par colonne
        stats = {col: {"neg": 0, "ident": 0, "vides": 0} for col in common_cols}
        total_cells = total_rows * nb_cols
        total_neg = 0
        total_vides = 0
        total_ident = 0

        # parcourir chaque colonne & ligne
        for col_idx, col in enumerate(common_cols):
            for row_idx in range(total_rows):
                val1 = merged.at[row_idx, f"{col}_1"] if f"{col}_1" in merged.columns else ""
                val2 = merged.at[row_idx, f"{col}_2"] if f"{col}_2" in merged.columns else ""

                is_val1_empty = (pd.isna(val1) or val1 == "")
                is_val2_empty = (pd.isna(val2) or val2 == "")

                # définir vide : les deux côtés vides
                if is_val1_empty and is_val2_empty:
                    stats[col]["vides"] += 1
                    total_vides += 1
                    # ne pas marquer en rouge (c'est vide)
                    continue

                # comparer
                if safe_compare(val1, val2):
                    # identique (et pas vide, sinon on l'aurait compté plus haut)
                    stats[col]["ident"] += 1
                    total_ident += 1
                else:
                    stats[col]["neg"] += 1
                    total_neg += 1
                    # colorer les deux cellules correspondantes dans l'excel de sortie
                    # trouver position colonne dans WS : header + 1, puis position 1-based des colonnes
                    # Le fichier excel contient d'abord merge_key, puis paires col_1, col_2 dans l'ordre des common_cols
                    # index: merge_key (col 1), ensuite pour col0 -> col1_idx=2, col2_idx=3, col1 of next col=4...
                    col1_idx = 2 + col_idx * 2
                    col2_idx = col1_idx + 1
                    excel_row = row_idx + 2  # because header row is row 1
                    try:
                        ws.cell(row=excel_row, column=col1_idx).fill = red_fill
                        ws.cell(row=excel_row, column=col2_idx).fill = red_fill
                    except:
                        pass

        # ajuster largeur
        auto_adjust_column_width(ws)

        # sauvegarder fichier
        if output_name and str(output_name).strip() != "":
            out_name = str(output_name).strip()
            if not out_name.lower().endswith(".xlsx"):
                out_name = out_name + ".xlsx"
        else:
            out_name = f"comparaison_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            temp_path = tmp.name

                # --- Préparer le camembert (fond sombre, pas de bord noir) ---
        red_cells = total_neg
        empty_cells = total_vides
        ident_cells = total_ident
        total_cells_safe = max(total_cells, 1)

        sizes = [red_cells, ident_cells, empty_cells]
        labels = ["Difference", "Identiques", "Vides"]
        colors = ['#d62728', '#1f77b4', '#2ca02c']  # rouge, bleu, vert

        # figure avec fond assorti au thème sombre pour éviter "barres/noires" autour
        fig, ax = plt.subplots(figsize=(6,6), dpi=100, facecolor='#121212')
        ax.set_facecolor('#121212')

        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            autopct=lambda pct: f"{pct:.1f}%\n({int(round(pct * total_cells_safe / 100))})" if total_cells>0 else "",
            startangle=90,
            colors=colors,
            wedgeprops=dict(width=0.6, edgecolor='#121212') , # enlever bord blanc/noir en mettant edgecolor = background
            textprops={'color': 'white'}  # étiquettes en blanc
        )

        # autotexts (pourcent + counts) : les rendre lisibles
        for txt in autotexts:
            txt.set_color('white')
            txt.set_fontsize(10)

        for t in texts:
            t.set_color('white')
            t.set_fontsize(11)

        # rendre le camembert bien centré et plein
        ax.axis('equal')
        ax.set_title("Proportion cellule en rouge", fontsize=14, color='white')

        # ajustements pour éviter rognage
        fig.subplots_adjust(left=0.02, right=0.98, top=0.92, bottom=0.02)

        # Préparer DataFrame de stats par colonne (triée)
        rows = []
        for col in common_cols:
            neg = stats[col]["neg"]
            ident = stats[col]["ident"]
            v = stats[col]["vides"]

            pct_error = (neg / total_rows * 100) if total_rows > 0 else 0

            rows.append({
                "Colonne": col,
                "Nb_negatifs": neg,
                "Nb_identiques": ident,
                "Nb_vides": v,
                "%_error": f"{round(pct_error,2)} %"
            })
        neg_df = pd.DataFrame(rows).sort_values(by="Nb_negatifs", ascending=False).reset_index(drop=True)
                # ---- Résumé global ----

        # ---- KPI Markdown professionnel ----

        total_cells_safe = max(total_cells, 1)

        pct_error = round((total_neg / total_cells_safe) * 100, 1)
        pct_ident = round((total_ident / total_cells_safe) * 100, 1)
        pct_empty = round((total_vides / total_cells_safe) * 100, 1)

        kpi_md = f"""
        <div style="display:flex; flex-direction:column; gap:12px; font-size:14px">

        <div style="padding:10px; border-radius:8px; background-color:#2b2b2b">
        <b style="color:#d62728">❌ Erreurs</b><br>
        <span style="font-size:20px">{total_neg}</span><br>
        {pct_error} %
        </div>

        <div style="padding:10px; border-radius:8px; background-color:#2b2b2b">
        <b style="color:#1f77b4">✅ Identiques</b><br>
        <span style="font-size:20px">{total_ident}</span><br>
        {pct_ident} %
        </div>

        <div style="padding:10px; border-radius:8px; background-color:#2b2b2b">
        <b style="color:#2ca02c">🟢 Vides</b><br>
        <span style="font-size:20px">{total_vides}</span><br>
        {pct_empty} %
        </div>

        <div style="padding:10px; border-radius:8px; background-color:#1a1a1a">
        <b>📊 Total cellules comparées</b><br>
        <span style="font-size:20px">{total_cells}</span>
        </div>

        </div>
        """


        return temp_path, "", fig, kpi_md, neg_df

    except Exception as e:
        return None, f"❌ Erreur : {e}", None, pd.DataFrame()

# --- Le reste de l'UI Gradio reste identique à ta version précédente ---
# (reprends l'interface Gradio que tu avais : inputs, preview, boutons, mapping outputs)
# Assure-toi que btn_compare retourne [output_file, error_msg, pie_chart, negatives_table]

# 🎛️ Interface Gradio
with gr.Blocks() as app:
    gr.Markdown("## 🔍 Comparaison de fichiers Excel — avec stats UI (non enregistrées dans l'Excel)")

    with gr.Row():
        file1 = gr.File(label="📁 Fichier Excel 1")
        file2 = gr.File(label="📁 Fichier Excel 2")

    with gr.Row():
        sheet1 = gr.Dropdown(label="Onglet Fichier 1", choices=[], value=None)
        sheet2 = gr.Dropdown(label="Onglet Fichier 2", choices=[], value=None)

    with gr.Row():
        header1 = gr.Number(label="Ligne des en-têtes Fichier 1", value=1)
        header2 = gr.Number(label="Ligne des en-têtes Fichier 2", value=1)

    with gr.Row():
        preview1 = gr.Dataframe(label="Aperçu Fichier 1")
        preview2 = gr.Dataframe(label="Aperçu Fichier 2")

    with gr.Row():
        btn_unique1 = gr.Button("🔎 Analyser Unique ID Fichier 1")
        btn_unique2 = gr.Button("🔎 Analyser Unique ID Fichier 2")

    with gr.Row():
        unique_result1 = gr.Dataframe(label="Top 5 colonnes les plus uniques - Fichier 1")
        unique_result2 = gr.Dataframe(label="Top 5 colonnes les plus uniques - Fichier 2")

    with gr.Row():
        btn_cols1 = gr.Button("🔃 Charger colonnes Fichier 1")
        dropdown_cols1 = gr.Dropdown(label="Colonne clé Fichier 1", choices=[], interactive=True)

    with gr.Row():
        btn_cols2 = gr.Button("🔃 Charger colonnes Fichier 2")
        dropdown_cols2 = gr.Dropdown(label="Colonne clé Fichier 2", choices=[], interactive=True)

    with gr.Row():
        output_name = gr.Textbox(label="Nom du fichier de sortie (optionnel)", placeholder="laisser vide pour nom par défaut (horodaté)")
    with gr.Row():
        btn_compare = gr.Button("📤 Comparer et exporter Excel")

    with gr.Row():
        output_file = gr.File(label="📄 Fichier Excel comparé")
        error_msg = gr.Textbox(label="Message d'erreur", interactive=False)

    # ---- Layout : camembert à gauche, KPI à droite, tableau full-width dessous ----
    with gr.Row():

        # ---- COLONNE GAUCHE (camembert) ----
        with gr.Column(scale=2):
            pie_chart = gr.Plot(label="")   # camembert seule

        # ---- COLONNE DROITE (KPI) ----
        with gr.Column(scale=1):
            kpi_markdown = gr.Markdown()    # KPI stylés à droite

    # ---- Ligne séparée, full-width, pour le tableau de détails ----
    with gr.Row():
        negatives_table = gr.Dataframe(
            label="Détail par colonne",
            interactive=False
        )
    # 📑 Fonctions auxiliaires
    def get_sheet_names(file):
        if file is None:
            return gr.update(choices=[], value=None)
        try:
            xls = pd.ExcelFile(file.name)
            sheets = xls.sheet_names
            default = sheets[0] if sheets else None
            return gr.update(choices=sheets, value=default)
        except:
            return gr.update(choices=[], value=None)

    def read_excel(file, sheet_name, header_row):
        if file is None or sheet_name is None or header_row is None:
            return pd.DataFrame()
        try:
            df = pd.read_excel(file.name, sheet_name=sheet_name, header=int(header_row)-1)
            # Nettoyer colonnes Unnamed pour affichage aperçu
            df = df.loc[:, filter_named_columns(df.columns)]
            return df.head(10)
        except:
            return pd.DataFrame()

    def get_columns(file, sheet_name, header_row):
        if file is None or sheet_name is None or header_row is None:
            return gr.update(choices=[], value=None)
        try:
            df = pd.read_excel(file.name, sheet_name=sheet_name, header=int(header_row)-1)
            cols = filter_named_columns(df.columns)
            return gr.update(choices=cols, value=cols[0] if cols else None)
        except:
            return gr.update(choices=[], value=None)

    # ⚙️ Connexion des événements Gradio
    file1.change(fn=get_sheet_names, inputs=file1, outputs=sheet1)
    file2.change(fn=get_sheet_names, inputs=file2, outputs=sheet2)

    sheet1.change(fn=read_excel, inputs=[file1, sheet1, header1], outputs=preview1)
    sheet2.change(fn=read_excel, inputs=[file2, sheet2, header2], outputs=preview2)

    btn_cols1.click(fn=get_columns, inputs=[file1, sheet1, header1], outputs=dropdown_cols1)
    btn_cols2.click(fn=get_columns, inputs=[file2, sheet2, header2], outputs=dropdown_cols2)

    btn_unique1.click(
        fn=compute_unique_ids,
        inputs=[file1, sheet1, header1],
        outputs=unique_result1
    )

    btn_unique2.click(
        fn=compute_unique_ids,
        inputs=[file2, sheet2, header2],
        outputs=unique_result2
    )

    # IMPORTANT: outputs order must match the function return order
    btn_compare.click(
        fn=compare_excels,
        inputs=[file1, sheet1, header1, dropdown_cols1, file2, sheet2, header2, dropdown_cols2, output_name],
        outputs=[output_file, error_msg, pie_chart, kpi_markdown, negatives_table]
    )

# 🚀 Lancement de l'app
app.launch(inbrowser=True)
